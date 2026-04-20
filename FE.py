import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
import numpy as np
import json
import re 
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

try:
    from openpyxl.cell.rich_text import CellRichText, TextBlock
    from openpyxl.cell.text import InlineFont
    HAS_RICH_TEXT = True
except ImportError:
    HAS_RICH_TEXT = False

class AdvancedFECalculatorApp:
    def __init__(self, master):
        self.master = master
        master.title("Faradaic Efficiency 計算器 (支援通氣模式切換)")
        master.geometry("1150x720") 

        # --- 歷史紀錄堆疊 (Undo 快照) ---
        self.history_stack = []

        # --- 模式與全域變數 ---
        self.mode_var = tk.StringVar(value="GDE") 
        self.last_mode = "GDE"
        
        # 共享與特定參數
        self.total_charge_var = tk.DoubleVar(value=100.0) 
        self.electrolyte_var = tk.StringVar(value='0.5 M KNO3 + 0.1 M KOH') 
        self.acid_vol_var = tk.DoubleVar(value=10.0)
        self.re_vol_var = tk.DoubleVar(value=50.0)
        
        # 💡 通氣模式開關：False = 氬氣 (Ar), True = 氮氣 (N2)
        self.gde_q_toggle_var = tk.BooleanVar(value=False)

        # --- 自訂公式變數 ---
        self.formula_hcell_fe = tk.StringVar(value="(Conc * 50 * Dilution * 1e-6 * n_e * F) / Q * 100")
        self.formula_gde_n = tk.StringVar(value="(C1 * V_acid + C2 * V_re) * Dilution")
        self.formula_gde_fe = tk.StringVar(value="(Total_n * 1e-6 * n_e * F) / Q * 100")
        
        # --- 啟動時先呼叫模式選擇視窗 ---
        self.show_startup_dialog()

    # --- 0. 歷史快照核心 (Undo 系統) ---
    def _get_current_state(self):
        state = {
            'global_params': {
                'mode': self.mode_var.get(),
                'total_coulomb': self.total_charge_var.get(),
                'electrolyte': self.electrolyte_var.get(),
                'acid_vol': self.acid_vol_var.get(),
                're_vol': self.re_vol_var.get(),
                'gde_q_toggle': self.gde_q_toggle_var.get(),
                'formula_hcell_fe': self.formula_hcell_fe.get(),
                'formula_gde_n': self.formula_gde_n.get(),
                'formula_gde_fe': self.formula_gde_fe.get()
            },
            'rows': []
        }
        for row in self.rows_data:
            row_data = {}
            for k, v in row['widgets'].items():
                if k != 'checked':
                    row_data[k] = v.get()
            state['rows'].append(row_data)
        return state

    def _save_state(self):
        self.history_stack.append(self._get_current_state())
        if len(self.history_stack) > 30: 
            self.history_stack.pop(0)

    def undo(self):
        if not self.history_stack:
            messagebox.showinfo("提示", "已經沒有上一步可以復原了！")
            return
        
        prev_state = self.history_stack.pop()
        target_mode = prev_state['global_params']['mode']
        gp = prev_state['global_params']
        
        self.total_charge_var.set(gp.get('total_coulomb', 100.0))
        self.electrolyte_var.set(gp.get('electrolyte', '0.5 M KNO3 + 0.1 M KOH'))
        self.acid_vol_var.set(gp.get('acid_vol', 10.0))
        self.re_vol_var.set(gp.get('re_vol', 50.0))
        self.gde_q_toggle_var.set(gp.get('gde_q_toggle', False))
        self.formula_hcell_fe.set(gp.get('formula_hcell_fe', "(Conc * 50 * Dilution * 1e-6 * n_e * F) / Q * 100"))
        self.formula_gde_n.set(gp.get('formula_gde_n', "(C1 * V_acid + C2 * V_re) * Dilution"))
        self.formula_gde_fe.set(gp.get('formula_gde_fe', "(Total_n * 1e-6 * n_e * F) / Q * 100"))

        if self.mode_var.get() != target_mode:
            self.mode_var.set(target_mode)
            self.last_mode = target_mode
            self.build_ui_for_mode(load_initial=False)
        else:
            self._clear_all_rows()
            # 重新繪製開關狀態
            if hasattr(self, 'draw_switch_func'):
                self.draw_switch_func()

        for row_data in prev_state.get('rows', []):
            self._insert_row_widgets(row_data)
            
        # 套用通氣模式的防呆限制
        self._apply_gas_mode_constraints()

    # --- 0.5 啟動模式選擇視窗 ---
    def show_startup_dialog(self):
        self.master.withdraw() 
        dialog = tk.Toplevel(self.master)
        dialog.title("啟動設定")
        dialog.resizable(False, False) 
        frame = ttk.Frame(dialog, padding=25)
        frame.pack(fill='both', expand=True)
        ttk.Label(frame, text="🧪  Faradaic Efficiency 數據計算機", font=('Arial', 14, 'bold'), foreground="darkblue").pack(pady=(0, 10))
        ttk.Label(frame, text="請選擇您要進行的實驗模式：", font=('Arial', 11)).pack(pady=(0, 20))
        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill='x')
        
        def select_mode(mode):
            self.mode_var.set(mode)
            self.last_mode = mode
            dialog.destroy()
            self.master.deiconify() 
            self.build_main_app()   

        style = ttk.Style()
        style.configure("Big.TButton", font=('Arial', 11, 'bold'), padding=10)

        btn_hcell = ttk.Button(btn_frame, text="H-cell 模式\n(單槽濃度計算)", style="Big.TButton", command=lambda: select_mode("H-cell"))
        btn_hcell.pack(side=tk.LEFT, padx=10, expand=True, fill='x')
        btn_gde = ttk.Button(btn_frame, text="GDE 模式\n(雙槽濃度計算)", style="Big.TButton", command=lambda: select_mode("GDE"))
        btn_gde.pack(side=tk.LEFT, padx=10, expand=True, fill='x')
        dialog.protocol("WM_DELETE_WINDOW", self.master.destroy)
        self.center_window(dialog)

    # --- 建立主程式 UI 框架 ---
    def build_main_app(self):
        self.create_top_bar() 
        self.main_container = ttk.Frame(self.master)
        self.main_container.pack(fill='both', expand=True)
        self.build_ui_for_mode() 
        
        self.master.bind_all('<MouseWheel>', self._on_mousewheel)
        self.master.bind_all('<Button-4>', self._on_mousewheel_linux) 
        self.master.bind_all('<Button-5>', self._on_mousewheel_linux) 

    # --- 1. 模式切換邏輯 ---
    def create_top_bar(self):
        top_frame = ttk.Frame(self.master)
        top_frame.pack(fill='x', padx=10, pady=5)
        ttk.Label(top_frame, text="🧪 當前實驗分析模式: ", font=('Arial', 11, 'bold'), foreground="darkblue").pack(side=tk.LEFT, padx=5)
        style = ttk.Style()
        style.configure("TRadiobutton", font=('Arial', 10, 'bold'))

        ttk.Radiobutton(top_frame, text="GDE 模式 (雙槽濃度與 FE 計算)", variable=self.mode_var, value="GDE", command=self.on_mode_change).pack(side=tk.LEFT, padx=15)
        ttk.Radiobutton(top_frame, text="H-cell 模式 (單一濃度 FE 計算)", variable=self.mode_var, value="H-cell", command=self.on_mode_change).pack(side=tk.LEFT, padx=15)
        ttk.Separator(self.master, orient='horizontal').pack(fill='x', padx=10, pady=2)

    def on_mode_change(self):
        new_mode = self.mode_var.get()
        if new_mode == self.last_mode: return
        if len(self.rows_data) > 0 and (len(self.rows_data) > 1 or self.rows_data[0]['widgets']['product'].get() != ""):
            if not messagebox.askyesno("切換模式警告", "切換模式將會清空當前表格中的所有輸入數據！\n\n確定要切換嗎？"):
                self.mode_var.set(self.last_mode)
                return
        self._save_state() 
        self.last_mode = new_mode
        self.build_ui_for_mode()

    def build_ui_for_mode(self, load_initial=True):
        for widget in self.main_container.winfo_children(): widget.destroy()
        self.rows_data = []
        self.current_row = 1
        self.row_counter = 1
        self.create_global_param_frame()
        self.create_table_frame()
        self.create_selection_frame()
        self.create_button_frame()
        if load_initial:
            self.load_initial_data()
            
        self._apply_gas_mode_constraints()

    # --- 2. 建立全域參數區 ---
    def create_global_param_frame(self):
        param_frame = ttk.LabelFrame(self.main_container, text="⚙️ 固定參數設定")
        param_frame.pack(padx=10, pady=5, fill='x')
        
        bg_color = ttk.Style().lookup('TFrame', 'background')
        if not bg_color: bg_color = "#F0F0F0"

        col_offset = 0

        # 💡 iOS 通氣模式開關繪製
        if self.mode_var.get() == "GDE":
            # 新增開關標籤
            self.lbl_gas_mode = ttk.Label(param_frame, text="Ar 氣", font=('Arial', 10, 'bold'), foreground="gray")
            self.lbl_gas_mode.grid(row=0, column=0, padx=(10, 2), pady=2, sticky='e')

            switch_canvas = tk.Canvas(param_frame, width=44, height=24, highlightthickness=0, bg=bg_color)
            switch_canvas.grid(row=0, column=1, padx=(0, 10), pady=2)
            
            def draw_switch():
                switch_canvas.delete("all")
                is_n2_mode = self.gde_q_toggle_var.get()
                
                if is_n2_mode:
                    color = "#4CD964" # 綠色
                    self.lbl_gas_mode.config(text="氮氣", foreground="#008000") # 深綠字
                else:
                    color = "#E5E5E5" # 灰色
                    self.lbl_gas_mode.config(text="氬氣", foreground="gray")

                switch_canvas.create_oval(2, 2, 22, 22, fill=color, outline=color)
                switch_canvas.create_oval(22, 2, 42, 22, fill=color, outline=color)
                switch_canvas.create_rectangle(12, 2, 32, 22, fill=color, outline=color)
                
                if is_n2_mode:
                    switch_canvas.create_oval(24, 4, 40, 20, fill="white", outline="")
                else:
                    switch_canvas.create_oval(4, 4, 20, 20, fill="white", outline="#D5D5D5")
                    
            self.draw_switch_func = draw_switch # 儲存參照讓 Undo 可以呼叫

            def on_toggle(event):
                self._save_state() 
                self.gde_q_toggle_var.set(not self.gde_q_toggle_var.get())
                draw_switch()
                self._apply_gas_mode_constraints() # 觸發防呆連動邏輯
                
            switch_canvas.bind("<Button-1>", on_toggle)
            draw_switch() 
            col_offset = 2

        ttk.Label(param_frame, text="總電量 Q (C):", font=('Arial', 10, 'bold')).grid(row=0, column=col_offset, padx=5, pady=2, sticky='w')
        ttk.Entry(param_frame, textvariable=self.total_charge_var, width=10).grid(row=0, column=col_offset+1, padx=5, pady=2, sticky='w')
        ttk.Label(param_frame, text="電解液:").grid(row=0, column=col_offset+2, padx=5, pady=2, sticky='w')
        ttk.Entry(param_frame, textvariable=self.electrolyte_var, width=25).grid(row=0, column=col_offset+3, padx=5, pady=2, sticky='w')
        
        if self.mode_var.get() == "GDE":
            ttk.Label(param_frame, text="Acid 側體積 (mL):").grid(row=0, column=col_offset+4, padx=5, pady=2, sticky='w')
            ttk.Entry(param_frame, textvariable=self.acid_vol_var, width=10).grid(row=0, column=col_offset+5, padx=5, pady=2, sticky='w')
            ttk.Label(param_frame, text="RE 側體積 (mL):").grid(row=0, column=col_offset+6, padx=5, pady=2, sticky='w')
            ttk.Entry(param_frame, textvariable=self.re_vol_var, width=10).grid(row=0, column=col_offset+7, padx=5, pady=2, sticky='w')

    # --- 3. 建立表格與輸入區 ---
    def create_table_frame(self):
        self.table_container = ttk.LabelFrame(self.main_container, text=f"📊 實驗數據輸入 ({self.mode_var.get()})")
        self.table_container.pack(padx=10, pady=5, fill='both', expand=True)
        self.table_canvas = tk.Canvas(self.table_container)
        self.table_canvas.pack(side="left", fill="both", expand=True)
        self.scrollbar = ttk.Scrollbar(self.table_container, orient="vertical", command=self.table_canvas.yview)
        self.scrollbar.pack(side="right", fill="y")
        self.table_canvas.configure(yscrollcommand=self.scrollbar.set)
        self.table_body = ttk.Frame(self.table_canvas)
        self.table_canvas.create_window((0, 0), window=self.table_body, anchor="nw")
        
        self.table_body.bind("<Configure>", lambda e: self.table_canvas.configure(scrollregion=self.table_canvas.bbox("all")))

        if self.mode_var.get() == "H-cell":
            self.columns = [
                ('No.', 35), ('☑', 25), ('Product', 60), ('Catalyst', 100),
                ('Loading (μl)', 80), ('V vs RHE', 80), ('稀釋倍率', 80), 
                ('Conc. (μmol)', 100), ('FE (%)', 100)
            ]
        else:
            self.columns = [
                ('No.', 35), ('☑', 25), ('Product', 60), ('Catalyst', 90),
                ('Loading (μl)', 75), ('V vs RHE', 75), ('稀釋倍率', 60), 
                ('Acid C1 (mM)', 80), ('RE C2 (mM)', 80), ('Total n (μmol)', 90), 
                ('FE (%)', 80)           
            ]
        
        for col_idx, (col_name, width) in enumerate(self.columns):
            lbl = ttk.Label(self.table_body, text=col_name, relief="raised", width=int(width/7), anchor='center', font=('Arial', 9, 'bold'))
            lbl.grid(row=0, column=col_idx, sticky="nsew")

    def _create_entry(self, row, col, initial_value, width):
        entry_var = tk.StringVar(value=str(initial_value))
        entry = ttk.Entry(self.table_body, textvariable=entry_var, width=width)
        entry.grid(row=row, column=col, padx=2, pady=2)
        return entry_var

    def _insert_row_widgets(self, initial_values):
        row_widgets = {}
        mode = self.mode_var.get()
        col = 0
        
        ttk.Label(self.table_body, text=str(self.row_counter), width=4, anchor='center').grid(row=self.current_row, column=col, padx=2, pady=2); col += 1
        check_var = tk.BooleanVar(value=initial_values.get('checked', False))
        ttk.Checkbutton(self.table_body, variable=check_var).grid(row=self.current_row, column=col, padx=2, pady=2)
        row_widgets['checked'] = check_var; col += 1
        
        product_var = tk.StringVar(value=initial_values.get('product', 'NH3'))
        product_combo = ttk.Combobox(self.table_body, textvariable=product_var, values=['NH3', 'NO2'], state="readonly", width=6)
        product_combo.grid(row=self.current_row, column=col, padx=2, pady=2)
        row_widgets['product'] = product_var
        row_widgets['product_combo_widget'] = product_combo # 記住元件，方便後續鎖死操作
        col += 1

        row_widgets['catalyst'] = self._create_entry(self.current_row, col, initial_values.get('catalyst', ''), width=12); col += 1
        row_widgets['volume_ul'] = self._create_entry(self.current_row, col, initial_values.get('volume_ul', 80.0), width=8); col += 1
        row_widgets['v_rhe'] = self._create_entry(self.current_row, col, initial_values.get('v_rhe', -1.4), width=8); col += 1
        row_widgets['dilution'] = self._create_entry(self.current_row, col, initial_values.get('dilution', '1'), width=8); col += 1

        if mode == "H-cell":
            row_widgets['conc'] = self._create_entry(self.current_row, col, initial_values.get('conc', 0.0), width=10); col += 1
            fe_var = tk.StringVar(value=initial_values.get('fe', "--"))
            ttk.Label(self.table_body, textvariable=fe_var, relief="groove", width=12, anchor='center', foreground="blue").grid(row=self.current_row, column=col, padx=2, pady=2)
            row_widgets['fe'] = fe_var
        else:
            row_widgets['acid_c1'] = self._create_entry(self.current_row, col, initial_values.get('acid_c1', 0.0), width=8); col += 1
            row_widgets['re_c2'] = self._create_entry(self.current_row, col, initial_values.get('re_c2', 0.0), width=8); col += 1
            total_var = tk.StringVar(value=initial_values.get('total_umol', "--"))
            ttk.Label(self.table_body, textvariable=total_var, relief="groove", width=10, anchor='center').grid(row=self.current_row, column=col, padx=2, pady=2)
            row_widgets['total_umol'] = total_var; col += 1
            fe_var = tk.StringVar(value=initial_values.get('fe', "--"))
            ttk.Label(self.table_body, textvariable=fe_var, relief="groove", width=10, anchor='center', foreground="blue").grid(row=self.current_row, column=col, padx=2, pady=2)
            row_widgets['fe'] = fe_var

        self.rows_data.append({'id': self.current_row, 'index_value': self.row_counter, 'widgets': row_widgets})
        self.current_row += 1
        self.row_counter += 1 

    # --- 💡 新增：通氣模式防呆邏輯 ---
    def _apply_gas_mode_constraints(self):
        """ 根據開關狀態，動態鎖死/解鎖 Product 欄位 """
        is_n2_mode = self.gde_q_toggle_var.get()
        for row in self.rows_data:
            combo = row['widgets'].get('product_combo_widget')
            prod_var = row['widgets'].get('product')
            if combo and prod_var:
                if is_n2_mode:
                    prod_var.set("NH3") # 強制設為 NH3
                    combo.config(state="disabled") # 鎖死不能改
                else:
                    combo.config(state="readonly") # 恢復可選

    # --- 5. 自訂公式編輯器 ---
    def open_formula_editor(self):
        popup = tk.Toplevel(self.master)
        popup.title("⚙️ 自訂計算公式設定")
        self.center_window(popup)
        ttk.Label(popup, text="提示：請使用標準數學符號 (+, -, *, /, **)，系統會自動帶入變數進行運算。", foreground="gray").pack(pady=5)
        lf_h = ttk.LabelFrame(popup, text="H-cell 模式: FE (%) 計算公式")
        lf_h.pack(fill='x', padx=15, pady=5)
        ttk.Label(lf_h, text="可用變數: Conc (濃度), Dilution (稀釋), Q (總電量), n_e (轉移電子數), F (96485)").pack(anchor='w', padx=5)
        ttk.Entry(lf_h, textvariable=self.formula_hcell_fe, width=70).pack(padx=5, pady=5)
        lf_g1 = ttk.LabelFrame(popup, text="GDE 模式: Total n (μmol) 計算公式")
        lf_g1.pack(fill='x', padx=15, pady=5)
        ttk.Label(lf_g1, text="可用變數: C1 (Acid濃度), C2 (RE濃度), V_acid (Acid體積), V_re (RE體積), Dilution").pack(anchor='w', padx=5)
        ttk.Entry(lf_g1, textvariable=self.formula_gde_n, width=70).pack(padx=5, pady=5)
        lf_g2 = ttk.LabelFrame(popup, text="GDE 模式: FE (%) 計算公式")
        lf_g2.pack(fill='x', padx=15, pady=5)
        ttk.Label(lf_g2, text="可用變數: Total_n (算出的總量), Q (總電量), n_e (轉移電子數), F (96485)").pack(anchor='w', padx=5)
        ttk.Entry(lf_g2, textvariable=self.formula_gde_fe, width=70).pack(padx=5, pady=5)

        def reset_defaults():
            self.formula_hcell_fe.set("(Conc * 50 * Dilution * 1e-6 * n_e * F) / Q * 100")
            self.formula_gde_n.set("(C1 * V_acid + C2 * V_re) * Dilution")
            self.formula_gde_fe.set("(Total_n * 1e-6 * n_e * F) / Q * 100")
            
        def apply_and_recalculate():
            popup.destroy()
            self.recalculate_fe()

        btn_frame = ttk.Frame(popup)
        btn_frame.pack(pady=10)
        ttk.Button(btn_frame, text="🔄 恢復預設", command=reset_defaults).pack(side=tk.LEFT, padx=10)
        ttk.Button(btn_frame, text="✔ 確定", command=apply_and_recalculate).pack(side=tk.LEFT, padx=10)

    # --- 排版與輔助工具 ---
    def load_initial_data(self):
        if self.mode_var.get() == "H-cell":
            initial_vals = {'catalyst': '', 'v_rhe': -1.4, 'product': 'NH3', 'conc': 0.0, 'dilution': '1', 'volume_ul': 80.0}
        else:
            initial_vals = {'catalyst': '', 'v_rhe': -1.4, 'product': 'NH3', 'acid_c1': 0.0, 're_c2': 0.0, 'dilution': '1', 'volume_ul': 80.0}
        self.add_new_row(count=1, initial_values=initial_vals)

    def _get_data_for_sort_or_export(self):
        extracted_data = []
        for row in self.rows_data:
            current_values = {}
            for key, var in row['widgets'].items():
                if key not in ['checked', 'product_combo_widget']: current_values[key] = var.get()
            current_values['index_value'] = row['index_value']
            current_values['checked'] = row['widgets']['checked'].get()
            extracted_data.append(current_values)
        return extracted_data

    def sort_rows_by_product(self): 
        self._save_state() 
        data_for_sorting = self._get_data_for_sort_or_export()
        def get_sort_key(data_dict):
            try: volume_sort = float(data_dict['volume_ul'])
            except ValueError: volume_sort = float('inf') 
            try: v_sort = float(data_dict['v_rhe'])
            except ValueError: v_sort = float('inf')
            return (0 if data_dict['product'] == 'NH3' else 1, data_dict['catalyst'], volume_sort, v_sort) 
        data_for_sorting.sort(key=get_sort_key)
        self._clear_all_rows()
        for data in data_for_sorting: self._insert_row_widgets(data)
        self._apply_gas_mode_constraints() # 重繪後補上鎖定狀態
        
    def create_selection_frame(self):
        select_frame = ttk.LabelFrame(self.main_container, text="🖱️ 選取操作 (請直接勾選表格前方的 ☑ 方塊)")
        select_frame.pack(padx=10, pady=5, fill='x')
        ttk.Button(select_frame, text="✔ 全選所有行", command=self.select_all_rows, style="Accent.TButton").pack(side=tk.LEFT, padx=10, pady=5)
        ttk.Button(select_frame, text="清除所有選取", command=self.clear_selection).pack(side=tk.LEFT, padx=5)

    def select_all_rows(self):
        for row in self.rows_data: row['widgets']['checked'].set(True)

    def clear_selection(self):
        for row in self.rows_data: row['widgets']['checked'].set(False)

    def _get_selected_rows(self):
        return [row for row in self.rows_data if row['widgets']['checked'].get()]

    def add_new_row(self, count=1, initial_values=None):
        if initial_values is None: initial_values = {}
        for _ in range(count): self._insert_row_widgets(initial_values)
        self._apply_gas_mode_constraints()
            
    def prompt_for_new_rows(self):
        popup = tk.Toplevel(self.master)
        popup.title("新增行數")
        self.center_window(popup)
        ttk.Label(popup, text="請輸入您要新增的行數：").pack(padx=20, pady=10)
        count_var = tk.StringVar(value="1")
        entry = ttk.Entry(popup, textvariable=count_var, width=10); entry.pack(padx=20, pady=5); entry.focus_set(); entry.selection_range(0, tk.END)
        def apply_add_rows(event=None):
            try:
                count = int(count_var.get())
                if count <= 0: raise ValueError
                self._save_state() 
                self.add_new_row(count); popup.destroy()
            except ValueError: messagebox.showerror("錯誤", "請輸入有效的正整數")
        ttk.Button(popup, text="確認", command=apply_add_rows).pack(pady=10)
        popup.bind('<Return>', apply_add_rows)

    def duplicate_selected_rows(self):
        selected_rows = self._get_selected_rows() 
        if not selected_rows: messagebox.showinfo("提示", "請先勾選要複製的行！"); return
        self._save_state() 
        for row in selected_rows:
            data = {k: v.get() for k, v in row['widgets'].items() if k not in ['checked', 'product_combo_widget']}
            self._insert_row_widgets(data)
        self._apply_gas_mode_constraints()

    def remove_selected_rows(self):
        selected_rows = self._get_selected_rows()
        if not selected_rows: messagebox.showinfo("提示", "請先勾選要刪除的行！"); return
        self._save_state() 
        selected_ids = [row['id'] for row in selected_rows]
        rows_to_keep = [row for row in self.rows_data if row['id'] not in selected_ids]
        deleted_count = len(self.rows_data) - len(rows_to_keep)
        self._clear_all_rows()
        for row in rows_to_keep: 
            data = {k: v.get() for k, v in row['widgets'].items() if k not in ['checked', 'product_combo_widget']}
            self._insert_row_widgets(data)
        self.clear_selection() 
        self._apply_gas_mode_constraints()
        messagebox.showinfo("完成", f"已成功刪除 {deleted_count} 行數據。")

    def bulk_edit_selected_rows(self):
        selected_rows = self._get_selected_rows() 
        if not selected_rows: messagebox.showinfo("提示", "請先勾選要編輯的行！"); return
        popup = tk.Toplevel(self.master)
        popup.title(f"批量編輯 ({len(selected_rows)} 行)")
        self.center_window(popup)
        main_frame = ttk.Frame(popup, padding="10")
        main_frame.pack(fill='both', expand=True)
        edit_vars = {}
        fields = [
            ("Product", tk.StringVar, ['NH3', 'NO2'], 'combobox'),
            ("Catalyst", tk.StringVar, None, 'entry'),
            ("Loading (μl)", tk.StringVar, None, 'entry'), 
            ("V vs RHE", tk.StringVar, None, 'entry'),
            ("稀釋倍率", tk.StringVar, None, 'entry')
        ]
        if self.mode_var.get() == "H-cell":
            fields.append(("Conc. (μmol)", tk.StringVar, None, 'entry'))
        else:
            fields.extend([("Acid C1 (mM)", tk.StringVar, None, 'entry'), ("RE C2 (mM)", tk.StringVar, None, 'entry')])
        
        ttk.Label(main_frame, text="輸入新值，留空則不修改：", font=('Arial', 10, 'bold')).grid(row=0, column=0, columnspan=2, pady=5)
        for i, (label_text, var_type, values, widget_type) in enumerate(fields, 1):
            ttk.Label(main_frame, text=f"{label_text}:").grid(row=i, column=0, padx=5, pady=2, sticky='w')
            var = var_type(value=""); edit_vars[label_text] = var
            if widget_type == 'entry': ttk.Entry(main_frame, textvariable=var, width=20).grid(row=i, column=1, padx=5, pady=2)
            else: 
                combo = ttk.Combobox(main_frame, textvariable=var, values=values, width=18)
                combo.grid(row=i, column=1, padx=5, pady=2)
                # 防呆：如果現在是 N2 模式，批量編輯的 Product 欄位也要鎖死
                if label_text == "Product" and self.gde_q_toggle_var.get() == True:
                    var.set("NH3")
                    combo.config(state="disabled")

        def apply_edit():
            mapping = {"Product": 'product', "Catalyst": 'catalyst', "Loading (μl)": 'volume_ul', "V vs RHE": 'v_rhe', "稀釋倍率": 'dilution', "Conc. (μmol)": 'conc', "Acid C1 (mM)": 'acid_c1', "RE C2 (mM)": 're_c2'} 
            updates = {mapping[k]: v.get().strip() for k, v in edit_vars.items() if v.get().strip()}
            if updates:
                self._save_state() 
                for row in selected_rows:
                    for k, v in updates.items():
                        if k in row['widgets'] and k != 'product_combo_widget': 
                            row['widgets'][k].set(v)
                messagebox.showinfo("完成", "更新成功"); popup.destroy()
            else: messagebox.showwarning("提示", "未輸入任何修改")
        ttk.Button(main_frame, text="應用修改", command=apply_edit).grid(row=len(fields)+1, column=0, columnspan=2, pady=10)

    # --- 系統控制 ---
    def _on_mousewheel(self, event):
        if event.widget.winfo_toplevel() != self.master: return
        if hasattr(self, 'table_canvas') and self.table_canvas.winfo_exists(): 
            if self.table_body.winfo_height() > self.table_canvas.winfo_height():
                self.table_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            
    def _on_mousewheel_linux(self, event):
        if event.widget.winfo_toplevel() != self.master: return
        if hasattr(self, 'table_canvas') and self.table_canvas.winfo_exists():
            if self.table_body.winfo_height() > self.table_canvas.winfo_height():
                if event.num == 4: self.table_canvas.yview_scroll(-1, "units")
                elif event.num == 5: self.table_canvas.yview_scroll(1, "units")
            
    def center_window(self, win):
        win.update_idletasks()
        x = (win.winfo_screenwidth() // 2) - (win.winfo_reqwidth() // 2)
        y = (win.winfo_screenheight() // 2) - (win.winfo_reqheight() // 2)
        win.geometry(f'+{x}+{y}')

    def _clear_all_rows(self):
        if hasattr(self, 'table_body') and self.table_body.winfo_exists():
            for child in self.table_body.winfo_children():
                if int(child.grid_info().get("row", 0)) > 0: child.destroy()
        self.rows_data = []; self.current_row = 1; self.row_counter = 1
        
        if hasattr(self, 'table_canvas') and self.table_canvas.winfo_exists():
            self.table_canvas.yview_moveto(0)

    # --- UI 雙排按鈕佈局 ---
    def create_button_frame(self):
        button_container = ttk.Frame(self.main_container)
        button_container.pack(pady=10, fill='x')

        row1_frame = ttk.Frame(button_container)
        row1_frame.pack(pady=5)
        row2_frame = ttk.Frame(button_container)
        row2_frame.pack(pady=5)
        style = ttk.Style()
        style.configure("Accent.TButton", font=('Arial', 10, 'bold'))

        # 第一排：檔案與全域操作
        ttk.Button(row1_frame, text="↩️ 復原 (Undo)", command=self.undo).pack(side=tk.LEFT, padx=5)
        ttk.Separator(row1_frame, orient='vertical').pack(side=tk.LEFT, fill='y', padx=10)

        ttk.Button(row1_frame, text="💾 儲存 JSON", command=self.save_data).pack(side=tk.LEFT, padx=5)
        ttk.Button(row1_frame, text="📂 讀取 JSON", command=self.load_data).pack(side=tk.LEFT, padx=5)
        ttk.Separator(row1_frame, orient='vertical').pack(side=tk.LEFT, fill='y', padx=10)
        
        ttk.Button(row1_frame, text="📥 匯出 Excel", command=self.export_data).pack(side=tk.LEFT, padx=5)
        ttk.Separator(row1_frame, orient='vertical').pack(side=tk.LEFT, fill='y', padx=10)
        
        ttk.Button(row1_frame, text="🔢 公式設定", command=self.open_formula_editor).pack(side=tk.LEFT, padx=5)
        ttk.Button(row1_frame, text="🔄 計算 FE", style="Accent.TButton", command=self.recalculate_fe).pack(side=tk.LEFT, padx=10)

        # 第二排：表格與資料列操作
        ttk.Button(row2_frame, text="➕ 新增", command=self.prompt_for_new_rows).pack(side=tk.LEFT, padx=5)
        ttk.Button(row2_frame, text="📋 複製", command=self.duplicate_selected_rows).pack(side=tk.LEFT, padx=5)
        ttk.Button(row2_frame, text="❌ 刪除", command=self.remove_selected_rows).pack(side=tk.LEFT, padx=5)
        ttk.Separator(row2_frame, orient='vertical').pack(side=tk.LEFT, fill='y', padx=10)
        
        ttk.Button(row2_frame, text="⚙️ 批量編輯", command=self.bulk_edit_selected_rows).pack(side=tk.LEFT, padx=5)
        ttk.Separator(row2_frame, orient='vertical').pack(side=tk.LEFT, fill='y', padx=10)
        
        ttk.Button(row2_frame, text="🔀 排序產物", command=self.sort_rows_by_product).pack(side=tk.LEFT, padx=5)

    # --- 儲存與讀取 ---
    def save_data(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")], title="儲存數據")
        if not file_path: return
        data_to_save = {
            'global_params': {
                'mode': self.mode_var.get(), 'total_coulomb': self.total_charge_var.get(),
                'electrolyte': self.electrolyte_var.get(), 'acid_vol': self.acid_vol_var.get(),
                're_vol': self.re_vol_var.get(), 'gde_q_toggle': self.gde_q_toggle_var.get(),
                'formula_hcell_fe': self.formula_hcell_fe.get(),
                'formula_gde_n': self.formula_gde_n.get(), 'formula_gde_fe': self.formula_gde_fe.get()
            },
            'rows': [{k: v.get() for k, v in row['widgets'].items() if k not in ['checked', 'product_combo_widget']} for row in self.rows_data]
        }
        try:
            with open(file_path, 'w', encoding='utf-8') as f: json.dump(data_to_save, f, ensure_ascii=False, indent=4)
            messagebox.showinfo("成功", f"數據與公式已儲存至 {file_path}")
        except Exception as e: messagebox.showerror("錯誤", f"儲存錯誤: {e}")

    def load_data(self):
        file_path = filedialog.askopenfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if not file_path: return
        try:
            with open(file_path, 'r', encoding='utf-8') as f: loaded_data = json.load(f)
            
            self._save_state() 
            
            params = loaded_data.get('global_params', {})
            if 'mode' in params: loaded_mode = params['mode']
            else:
                rows = loaded_data.get('rows', [])
                loaded_mode = 'H-cell' if len(rows) > 0 and 'conc' in rows[0] else 'GDE'
            
            self.total_charge_var.set(params.get('total_coulomb', 100.0))
            self.electrolyte_var.set(params.get('electrolyte', '0.5 M KNO3 + 0.1 M KOH'))
            self.acid_vol_var.set(params.get('acid_vol', 10.0))
            self.re_vol_var.set(params.get('re_vol', 50.0))
            self.gde_q_toggle_var.set(params.get('gde_q_toggle', False))

            if 'formula_hcell_fe' in params: self.formula_hcell_fe.set(params['formula_hcell_fe'])
            if 'formula_gde_n' in params: self.formula_gde_n.set(params['formula_gde_n'])
            if 'formula_gde_fe' in params: self.formula_gde_fe.set(params['formula_gde_fe'])

            self.mode_var.set(loaded_mode)
            self.last_mode = loaded_mode
            self.build_ui_for_mode(load_initial=False)

            self._clear_all_rows()
            for data in loaded_data.get('rows', []): self._insert_row_widgets(data)
            
            self._apply_gas_mode_constraints() # 讀取後套用防呆
            
            messagebox.showinfo("成功", f"數據載入成功！已切換至 {loaded_mode} 模式。")
        except Exception as e: messagebox.showerror("錯誤", f"載入錯誤: {e}")

    # --- 💡 核心計算 (新增 N2 模式電子數轉換) ---
    def recalculate_fe(self):
        self._save_state() 
        mode = self.mode_var.get()
        F_const = 96485  
        try:
            Q = float(self.total_charge_var.get())
            if Q <= 0: raise ValueError
            if mode == "GDE":
                vol_acid, vol_re = float(self.acid_vol_var.get()), float(self.re_vol_var.get())
                if any(val <= 0 for val in [vol_acid, vol_re]): raise ValueError
        except ValueError:
            messagebox.showerror("錯誤", "固定參數必須是有效正數"); return

        is_n2_mode = self.gde_q_toggle_var.get()

        for row in self.rows_data:
            w = row['widgets']
            try:
                prod = w['product'].get() 
                
                # 💡 判斷電子數：如果是 N2 模式且產物是 NH3，電子數為 6；否則按照正常規則
                if is_n2_mode and prod == 'NH3':
                    n_e = 6
                else:
                    n_e = 8 if prod == 'NH3' else (2 if prod == 'NO2' else np.nan)
                    
                dilution = float(w['dilution'].get())
                
                if mode == "H-cell":
                    conc = float(w['conc'].get())
                    if not np.isnan(n_e):
                        env = {'Conc': conc, 'Dilution': dilution, 'Q': Q, 'n_e': n_e, 'F': F_const}
                        fe_val = eval(self.formula_hcell_fe.get(), {"__builtins__": {}}, env)
                        w['fe'].set(f"{fe_val:.2f}") 
                    else:
                        w['fe'].set("Error")
                else:
                    c1, c2 = float(w['acid_c1'].get()), float(w['re_c2'].get())
                    env_n = {'C1': c1, 'C2': c2, 'V_acid': vol_acid, 'V_re': vol_re, 'Dilution': dilution}
                    total_n = eval(self.formula_gde_n.get(), {"__builtins__": {}}, env_n)
                    w['total_umol'].set(f"{total_n:.3f}")
                    
                    if not np.isnan(n_e):
                        env_fe = {'Total_n': total_n, 'Q': Q, 'n_e': n_e, 'F': F_const}
                        fe_val = eval(self.formula_gde_fe.get(), {"__builtins__": {}}, env_fe)
                        w['fe'].set(f"{fe_val:.2f}") 
                    else: 
                        w['fe'].set("Error")

            except Exception as e:
                if mode == "GDE": w['total_umol'].set("--")
                w['fe'].set("Err")

    # --- Excel 動態匯出 ---
    def _create_rich_text_with_subscript(self, text):
        if text is None or text == "": 
            return ""
            
        if isinstance(text, (int, float)) or type(text).__module__ == 'numpy':
            if pd.isna(text): return ""
            return text
            
        text_str = str(text)
        
        try:
            f_val = float(text_str)
            if np.isnan(f_val): return ""
            return int(f_val) if f_val.is_integer() else f_val
        except ValueError:
            pass
            
        subscript_map = str.maketrans("0123456789", "₀₁₂₃₄₅₆₇₈₉")
        
        def replace_sub(match):
            return match.group(1) + match.group(2).translate(subscript_map)
            
        return re.sub(r'([a-zA-Z])(\d+)', replace_sub, text_str)

    def export_data(self):
        self.recalculate_fe() 
        mode = self.mode_var.get()
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path: return

        gas_mode_str = "N2_Gas" if self.gde_q_toggle_var.get() else "Ar_Gas"

        export_list = []
        for row in self.rows_data:
            w = row['widgets']
            if mode == "H-cell":
                export_list.append({
                    'cell': f"H-cell({w['product'].get()})", 'electrolyte': self.electrolyte_var.get(),
                    'Total Coulomb (Q)': self.total_charge_var.get(), 'Product Type': w['product'].get(),
                    'catalyst': w['catalyst'].get(), 'Loading (μl)': w['volume_ul'].get(), '稀釋倍率': w['dilution'].get(), 
                    'V vs RHE': w['v_rhe'].get(), 'Total Concentration (μmol)': w['conc'].get(), 'Faradaic Efficiency (%)': w['fe'].get()
                })
            else:
                export_list.append({
                    'cell': f"GDE_{gas_mode_str}({w['product'].get()})", 'electrolyte': self.electrolyte_var.get(),
                    'Total Coulomb (Q)': self.total_charge_var.get(), 'Product Type': w['product'].get(),
                    'catalyst': w['catalyst'].get(), 'Loading (μl)': w['volume_ul'].get(), '稀釋倍率': w['dilution'].get(), 
                    'V vs RHE': w['v_rhe'].get(), 'Acid C1 (mM)': w['acid_c1'].get(), 'RE C2 (mM)': w['re_c2'].get(),
                    'Total Concentration (μmol)': w['total_umol'].get(), 'Faradaic Efficiency (%)': w['fe'].get()
                })
        
        if not export_list:
            messagebox.showwarning("提示", "表格中沒有數據可以匯出！")
            return
            
        # 1. 將資料依照 Product Type 分組
        grouped_data = {}
        for item in export_list:
            prod = item['Product Type']
            if prod not in grouped_data:
                grouped_data[prod] = []
            grouped_data[prod].append(item)

        wb = Workbook()  
        ws = wb.active
        ws.title = "FE_Data"
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        current_excel_row = 1 # 追蹤目前寫到 Excel 的第幾行

        # 2. 針對每一個產物區塊寫入同一個 Sheet
        for prod, data_list in grouped_data.items():
            df = pd.DataFrame(data_list)
            ordered_columns = list(data_list[0].keys())
            
            # 寫入標題列
            for c_idx, col_name in enumerate(ordered_columns):
                cell = ws.cell(row=current_excel_row, column=c_idx+1, value=self._create_rich_text_with_subscript(col_name))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 寫入數據列
            start_data_row = current_excel_row + 1
            for r_idx, row_data in enumerate(df[ordered_columns].values.tolist()):
                actual_row = start_data_row + r_idx
                for c_idx, value in enumerate(row_data):
                    val = self._create_rich_text_with_subscript(value) 
                    cell = ws.cell(row=actual_row, column=c_idx+1, value=val)
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            end_data_row = start_data_row + len(data_list) - 1

            # 3. 獨立處理「當前這一個表格區塊」的合併儲存格
            merge_levels = [
                [1, 2, 3], # L1: Cell, Electrolyte, Q
                [4],       # L2: Product Type
                [5],       # L3: Catalyst
                [6],       # L4: Loading (μl)
                [7]        # L5: 稀釋倍率
            ]
            
            starts = {i: start_data_row for i in range(len(merge_levels))}
            vals = {i: None for i in range(len(merge_levels))}

            for r in range(start_data_row, end_data_row + 2):
                if r <= end_data_row:
                    row_vals = {i: tuple(str(ws.cell(r, c).value) for c in cols) for i, cols in enumerate(merge_levels)}
                else:
                    row_vals = {i: None for i in range(len(merge_levels))}

                for i in range(len(merge_levels)):
                    changed = any(row_vals[p] != vals[p] for p in range(i + 1))
                    if changed:
                        end_r = r - 1
                        start_r = starts[i]
                        if end_r > start_r:
                            for c in merge_levels[i]:
                                ws.merge_cells(start_row=start_r, end_row=end_r, start_column=c, end_column=c)
                        starts[i] = r
                        vals[i] = row_vals[i]

            # 4. 更新行數，準備寫入下一個表格（+2 代表跳過一行空白行作為間隔）
            current_excel_row = end_data_row + 2

        # 5. 統一自動調整所有欄寬
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = (max((len(str(cell.value)) if hasattr(cell.value, '__str__') else len(cell.value)) for cell in col if cell.value) + 2) * 1.1

        try:
            wb.save(file_path)
            messagebox.showinfo("成功", f"數據已成功匯出至:\n{file_path}\n\n(✨ 產物區塊已上下排開)")
        except Exception as e:
            messagebox.showerror("匯出失敗", f"儲存檔案時發生錯誤。請確保檔案未開啟。\n錯誤: {e}")

if __name__ == '__main__':
    root = tk.Tk()
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except: pass
    app = AdvancedFECalculatorApp(root)
    root.mainloop()