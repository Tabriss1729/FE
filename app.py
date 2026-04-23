import streamlit as st
import pandas as pd
import numpy as np
import json
import re
import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font

# --- 1. 頁面與常數設定 ---
st.set_page_config(page_title="Faradaic Efficiency 計算機", layout="wide")
st.title("⚡ Faradaic Efficiency 數據計算機")

F_const = 96485
today_str = datetime.date.today().strftime("%Y%m%d")

# --- 初始化 Session State ---
if 'editor_key' not in st.session_state: st.session_state.editor_key = 0
if 'loaded_file_id' not in st.session_state: st.session_state.loaded_file_id = ""

if 'mode' not in st.session_state: st.session_state.mode = "GDE (雙槽)"
if 'q_toggle' not in st.session_state: st.session_state.q_toggle = False
if 'custom_ne_toggle' not in st.session_state: st.session_state.custom_ne_toggle = False
if 'sidebar_ne' not in st.session_state: st.session_state.sidebar_ne = 8.0 
if 'total_q' not in st.session_state: st.session_state.total_q = 100.0
if 'electrolyte' not in st.session_state: st.session_state.electrolyte = "0.5 M KNO3 + 0.1 M KOH" 
if 'acid_vol' not in st.session_state: st.session_state.acid_vol = 10.0
if 're_vol' not in st.session_state: st.session_state.re_vol = 50.0

if 'hcell_formula' not in st.session_state: st.session_state.hcell_formula = "(Conc * 50 * Dilution * 1e-6 * n_e * F) / Q * 100"
if 'gde_n_formula' not in st.session_state: st.session_state.gde_n_formula = "(C1 * V_acid + C2 * V_re) * Dilution"
if 'gde_fe_formula' not in st.session_state: st.session_state.gde_fe_formula = "(Total_n * 1e-6 * n_e * F) / Q * 100"

if 'hcell_data' not in st.session_state:
    st.session_state.hcell_data = pd.DataFrame({
        "選取": pd.Series(dtype='bool'),
        "Product": pd.Series(dtype='str'), "n_e": pd.Series(dtype='float'), "Catalyst": pd.Series(dtype='str'), "Loading (μl)": pd.Series(dtype='float'),
        "V vs RHE": pd.Series(dtype='float'), "Dilution Factor": pd.Series(dtype='float'), "Conc. (μmol)": pd.Series(dtype='float')
    })
if 'gde_data' not in st.session_state:
    st.session_state.gde_data = pd.DataFrame({
        "選取": pd.Series(dtype='bool'),
        "Product": pd.Series(dtype='str'), "n_e": pd.Series(dtype='float'), "Catalyst": pd.Series(dtype='str'), "Loading (μl)": pd.Series(dtype='float'),
        "V vs RHE": pd.Series(dtype='float'), "Dilution Factor": pd.Series(dtype='float'), "Acid C1 (mM)": pd.Series(dtype='float'), "RE C2 (mM)": pd.Series(dtype='float')
    })

# 相容性防呆
for df_name in ['hcell_data', 'gde_data']:
    if '選取' not in st.session_state[df_name].columns: 
        st.session_state[df_name].insert(0, '選取', False)
    if 'n_e' not in st.session_state[df_name].columns:
        st.session_state[df_name].insert(2, 'n_e', np.nan)

# JSON 讀取中繼處理
if 'pending_json_data' in st.session_state:
    data = st.session_state.pending_json_data
    gp = data.get('global_params', {})
    
    if 'mode' in gp: st.session_state.mode = "H-cell (單槽)" if gp['mode'] == "H-cell" else "GDE (雙槽)"
    st.session_state.q_toggle = gp.get('gde_q_toggle', False)
    st.session_state.custom_ne_toggle = gp.get('custom_ne_toggle', False)
    st.session_state.sidebar_ne = gp.get('sidebar_ne', 8.0)
    st.session_state.total_q = gp.get('total_coulomb', 100.0)
    st.session_state.electrolyte = gp.get('electrolyte', "0.5 M KNO3 + 0.1 M KOH")
    st.session_state.acid_vol = gp.get('acid_vol', 10.0)
    st.session_state.re_vol = gp.get('re_vol', 50.0)
    
    st.session_state.hcell_formula = gp.get('hcell_formula', "(Conc * 50 * Dilution * 1e-6 * n_e * F) / Q * 100")
    st.session_state.gde_n_formula = gp.get('gde_n_formula', "(C1 * V_acid + C2 * V_re) * Dilution")
    st.session_state.gde_fe_formula = gp.get('gde_fe_formula', "(Total_n * 1e-6 * n_e * F) / Q * 100")
    
    rows = data.get('rows', [])
    if rows:
        df = pd.DataFrame(rows)
        if '選取' not in df.columns: df['選取'] = False
        mapping = {'product': 'Product', 'catalyst': 'Catalyst', 'volume_ul': 'Loading (μl)', 'v_rhe': 'V vs RHE', 'dilution': 'Dilution Factor', '稀釋倍率': 'Dilution Factor', 'conc': 'Conc. (μmol)', 'acid_c1': 'Acid C1 (mM)', 're_c2': 'RE C2 (mM)'}
        df = df.rename(columns=mapping)
        if 'n_e' not in df.columns:
            df['n_e'] = df['Product'].apply(lambda p: 6.0 if (st.session_state.q_toggle and p == 'NH3') else (8.0 if p == 'NH3' else 2.0))
            
        if gp.get('mode') == "H-cell": st.session_state.hcell_data = df[[c for c in st.session_state.hcell_data.columns if c in df.columns]]
        else: st.session_state.gde_data = df[[c for c in st.session_state.gde_data.columns if c in df.columns]]
    
    st.session_state.editor_key += 1
    if 'res_df' in st.session_state: del st.session_state.res_df 
    del st.session_state.pending_json_data

def commit_edits():
    curr_mode = st.session_state.get('mode', "GDE (雙槽)")
    key = f"data_editor_{curr_mode}_{st.session_state.editor_key}"
    if key in st.session_state:
        state = st.session_state[key]
        df = st.session_state.hcell_data if "H-cell" in curr_mode else st.session_state.gde_data
        if "edited_rows" in state:
            for row_idx_str, changes in state["edited_rows"].items():
                r_idx = int(row_idx_str)
                if r_idx in df.index:
                    for col, val in changes.items():
                        df.loc[r_idx, col] = val

# --- 2. 側邊欄 ---
with st.sidebar:
    st.header("🧪 實驗參數")
    st.radio("實驗模式", ["H-cell (單槽)", "GDE (雙槽)"], key="mode")
    mode = st.session_state.mode
    
    if "GDE" in mode:
        st.toggle("通入氮氣 (N2 Mode)", key="q_toggle")
    is_n2_mode = st.session_state.q_toggle if "GDE" in mode else False 

    st.markdown("---")
    st.toggle("啟用自訂電子轉移數", key="custom_ne_toggle")
    col_q, col_ne = st.columns(2)
    with col_q:
        st.number_input("總電量 Q (C)", step=10.0, key="total_q")
    with col_ne:
        st.number_input("自訂 n_e", step=0.1, key="sidebar_ne", disabled=not st.session_state.custom_ne_toggle)
        
    total_q = st.session_state.total_q

    st.text_input("Electrolyte", key="electrolyte")
    electrolyte = st.session_state.electrolyte

    if "GDE" in mode:
        st.number_input("Acid 側體積 (mL)", key="acid_vol")
        st.number_input("RE 側體積 (mL)", key="re_vol")

    st.markdown("---")
    with st.expander("⚙️ 公式設定 (Formula)", expanded=False):
        st.markdown(f"""
        <details>
        <summary style="cursor: pointer; font-weight: bold; color: #4A90E2;">📖 變數說明與當前數值 (點擊展開)</summary>
        <ul style="margin-top: 8px; margin-bottom: 15px; padding-left: 20px; font-size: 0.9em; line-height: 1.6;">
            <li><code>Q</code> (總電量) = <b>{st.session_state.total_q}</b></li>
            <li><code>F</code> (法拉第常數) = <b>{F_const}</b></li>
            <li><code>V_acid</code> (Acid 側體積) = <b>{st.session_state.acid_vol}</b></li>
            <li><code>V_re</code> (RE 側體積) = <b>{st.session_state.re_vol}</b></li>
            <li><code>n_e</code> (轉移電子數) = <b>讀取表格各行</b></li>
            <li><code>Dilution</code> (稀釋倍率) = <b>讀取表格各行</b></li>
            <li><code>Conc</code> (莫耳濃度) = <b>讀取表格各行</b></li>
            <li><code>C1</code> (Acid 濃度) = <b>讀取表格各行</b></li>
            <li><code>C2</code> (RE 濃度) = <b>讀取表格各行</b></li>
            <li><code>Total_n</code> (總莫耳數) = <b>由 GDE 公式計算</b></li>
        </ul>
        </details>
        """, unsafe_allow_html=True)
        
        st.markdown("##### H-cell 公式")
        st.text_area("FE (%) =", height=68, label_visibility="collapsed", key="hcell_formula")
        st.divider()
        st.markdown("##### GDE 雙槽公式")
        st.text_input("Total n (μmol) =", key="gde_n_formula")
        st.text_input("FE (%) =", key="gde_fe_formula")
        
        if st.button("🔄 恢復預設公式", use_container_width=True):
            st.session_state.hcell_formula = "(Conc * 50 * Dilution * 1e-6 * n_e * F) / Q * 100"
            st.session_state.gde_n_formula = "(C1 * V_acid + C2 * V_re) * Dilution"
            st.session_state.gde_fe_formula = "(Total_n * 1e-6 * n_e * F) / Q * 100"
            st.rerun()

    st.markdown("---")
    st.header("💾 設定與檔案管理")
    with st.expander("📂 讀取 JSON 設定檔", expanded=True):
        uploaded_json = st.file_uploader("選擇檔案 (.json)", type="json", label_visibility="collapsed")
        if uploaded_json is not None:
            if st.session_state.loaded_file_id != uploaded_json.file_id:
                try:
                    file_content = uploaded_json.getvalue().decode("utf-8")
                    data = json.loads(file_content)
                    st.session_state.pending_json_data = data
                    st.session_state.loaded_file_id = uploaded_json.file_id
                    st.rerun()
                except Exception as e:
                    st.error(f"讀取失敗！詳細原因: {e}")

    st.markdown("##### 💾 儲存當前設定")
    custom_json_name = st.text_input("自訂 JSON 檔名", value="FE_Config", key="json_name_input")
    json_filename_final = f"{today_str}_{st.session_state.json_name_input}.json"

    save_data = {
        'global_params': {
            'mode': st.session_state.mode, 
            'total_coulomb': st.session_state.total_q, 
            'electrolyte': st.session_state.electrolyte, 
            'acid_vol': st.session_state.acid_vol, 
            're_vol': st.session_state.re_vol, 
            'gde_q_toggle': st.session_state.q_toggle,
            'custom_ne_toggle': st.session_state.custom_ne_toggle,
            'sidebar_ne': st.session_state.sidebar_ne,
            'hcell_formula': st.session_state.hcell_formula,
            'gde_n_formula': st.session_state.gde_n_formula,
            'gde_fe_formula': st.session_state.gde_fe_formula
        },
        'rows': st.session_state.hcell_data.to_dict('records') if "H-cell" in st.session_state.mode else st.session_state.gde_data.to_dict('records')
    }
    st.download_button("📥 儲存為 JSON", data=json.dumps(save_data, ensure_ascii=False, indent=4), file_name=json_filename_final, use_container_width=True)

# 處理手動編輯存檔
editor_key_str = f"data_editor_{mode}_{st.session_state.editor_key}"
target_df = st.session_state.hcell_data.copy() if "H-cell" in mode else st.session_state.gde_data.copy()

if editor_key_str in st.session_state:
    edits = st.session_state[editor_key_str].get("edited_rows", {})
    for row_idx_str, changes in edits.items():
        r_idx = int(row_idx_str)
        if r_idx in target_df.index:
            for col, val in changes.items():
                target_df.loc[r_idx, col] = val

# 動態設定 n_e 欄位的值
if st.session_state.custom_ne_toggle:
    target_df['n_e'] = st.session_state.sidebar_ne
else:
    target_df['n_e'] = target_df['Product'].apply(
        lambda p: 6.0 if (is_n2_mode and p == 'NH3') else (8.0 if p == 'NH3' else 2.0)
    )
if "H-cell" in mode: st.session_state.hcell_data = target_df
else: st.session_state.gde_data = target_df

# --- 3. 表格操作 ---
with st.expander("🛠️ 表格操作 (新增行數 / 批量修改 / 刪除)", expanded=False):
    st.markdown("##### ➕ 批量新增空行")
    col_add1, col_add2, _ = st.columns([1, 1, 3])
    with col_add1:
        add_count = st.number_input("輸入要新增的行數", min_value=1, max_value=50, value=1, step=1, label_visibility="collapsed", key="add_row_count")
    with col_add2:
        if st.button("確認新增"):
            new_rows = []
            default_ne = st.session_state.sidebar_ne if st.session_state.custom_ne_toggle else (6.0 if is_n2_mode else 8.0)
            for _ in range(st.session_state.add_row_count):
                if "H-cell" in mode:
                    new_rows.append({"選取": False, "Product": "NH3", "n_e": default_ne, "Catalyst": "", "Loading (μl)": None, "V vs RHE": None, "Dilution Factor": 1.0, "Conc. (μmol)": 0.0})
                else:
                    new_rows.append({"選取": False, "Product": "NH3", "n_e": default_ne, "Catalyst": "", "Loading (μl)": None, "V vs RHE": None, "Dilution Factor": 1.0, "Acid C1 (mM)": 0.0, "RE C2 (mM)": 0.0})
            
            new_df = pd.DataFrame(new_rows)
            target_df = pd.concat([target_df, new_df], ignore_index=True)
            if "H-cell" in mode: st.session_state.hcell_data = target_df
            else: st.session_state.gde_data = target_df
            st.session_state.editor_key += 1 
            st.rerun()

    st.divider()

    st.markdown("##### 🪄 批量修改與刪除 (請先在下方表格勾選 ☑)")
    cols = st.columns(5)
    with cols[0]:
        prod_options = ["(不修改)", "NH3"] if is_n2_mode else ["(不修改)", "NH3", "NO2"]
        b_prod = st.selectbox("Product", options=prod_options, key="b_prod")
    with cols[1]: b_cat = st.text_input("Catalyst", key="b_cat")
    with cols[2]: b_load = st.text_input("Loading (μl)", key="b_load")
    with cols[3]: b_vrhe = st.text_input("V vs RHE", key="b_vrhe")
    with cols[4]: b_dil = st.text_input("Dilution Factor", key="b_dil")
    
    col_btn1, col_btn2, _ = st.columns([2, 2, 4])
    with col_btn1:
        if st.button("🪄 套用修改至已選取行", use_container_width=True):
            try:
                mask = target_df["選取"] == True
                if not mask.any():
                    st.warning("⚠️ 請先在下方表格左側勾選 (☑) 您要修改的行！")
                else:
                    if st.session_state.get('b_prod', "(不修改)") != "(不修改)": target_df.loc[mask, "Product"] = st.session_state.b_prod
                    if st.session_state.get('b_cat'): target_df.loc[mask, "Catalyst"] = st.session_state.b_cat
                    if st.session_state.get('b_load'): target_df.loc[mask, "Loading (μl)"] = float(st.session_state.b_load)
                    if st.session_state.get('b_vrhe'): target_df.loc[mask, "V vs RHE"] = float(st.session_state.b_vrhe)
                    if st.session_state.get('b_dil'): target_df.loc[mask, "Dilution Factor"] = float(st.session_state.b_dil)
                    
                    target_df["選取"] = False
                    if "H-cell" in mode: st.session_state.hcell_data = target_df
                    else: st.session_state.gde_data = target_df
                    st.session_state.editor_key += 1
                    st.rerun()
            except ValueError:
                st.error("數值格式輸入錯誤！請確保數值欄位中輸入的是數字。")

    with col_btn2:
        if st.button("❌ 刪除已選取行", use_container_width=True):
            try:
                mask = target_df["選取"] == True
                if not mask.any():
                    st.warning("⚠️ 請先在下方表格左側勾選 (☑) 您要刪除的行！")
                else:
                    target_df = target_df[~mask].reset_index(drop=True)
                    if "H-cell" in mode: st.session_state.hcell_data = target_df
                    else: st.session_state.gde_data = target_df
                    st.session_state.editor_key += 1
                    st.rerun()
            except Exception as e:
                st.error(f"刪除失敗: {e}")

# --- 4. 數據表格 ---
st.subheader(f"📊 數據輸入 - {mode}")

col_sel1, col_sel2, _ = st.columns([1, 1, 8])
with col_sel1:
    if st.button("☑ 全選", use_container_width=True):
        target_df["選取"] = True
        if "H-cell" in mode: st.session_state.hcell_data = target_df
        else: st.session_state.gde_data = target_df
        st.session_state.editor_key += 1
        st.rerun()
with col_sel2:
    if st.button("☐ 全取消", use_container_width=True):
        target_df["選取"] = False
        if "H-cell" in mode: st.session_state.hcell_data = target_df
        else: st.session_state.gde_data = target_df
        st.session_state.editor_key += 1
        st.rerun()

base_render_df = st.session_state.hcell_data if "H-cell" in mode else st.session_state.gde_data

col_cfg = {
    "選取": st.column_config.CheckboxColumn("☑ 選取", default=False, width="small"),
    "Product": st.column_config.SelectboxColumn("Product", options=["NH3"] if is_n2_mode else ["NH3", "NO2"], required=True)
}
if st.session_state.custom_ne_toggle:
    col_cfg["n_e"] = st.column_config.NumberColumn("n_e", required=True)
else:
    col_cfg["n_e"] = None 

edited_df = st.data_editor(
    base_render_df,
    key=editor_key_str,
    num_rows="fixed",
    use_container_width=True,
    hide_index=True,
    on_change=commit_edits,
    column_config=col_cfg
)

# --- 5. 計算與匯出設定 ---
st.divider()
st.subheader("📥 計算與匯出")

if "H-cell" in mode:
    default_excel_name = "FE_Result_H-cell"
else:
    gas_str = "N2_Gas" if is_n2_mode else "Ar_Gas"
    default_excel_name = f"FE_Result_GDE_{gas_str}"

if st.button("🔄 開始計算 FE", type="primary"):
    res_df = edited_df.copy()
    
    if "H-cell" in mode:
        res_df["Conc. (μmol)"] = res_df["Conc. (μmol)"].fillna(0.0)
    else:
        res_df["Acid C1 (mM)"] = res_df["Acid C1 (mM)"].fillna(0.0)
        res_df["RE C2 (mM)"] = res_df["RE C2 (mM)"].fillna(0.0)

    fe_res, tn_res = [], []
    for _, row in res_df.iterrows():
        try:
            prod = row["Product"]
            
            row_ne = row.get("n_e")
            if pd.isna(row_ne) or row_ne == "":
                n_e = 6 if (is_n2_mode and prod == 'NH3') else (8 if prod == 'NH3' else (2 if prod == 'NO2' else np.nan))
            else:
                n_e = float(row_ne)
                
            dil = float(row["Dilution Factor"]) if pd.notna(row["Dilution Factor"]) else 1.0 
            
            if "H-cell" in mode:
                env = {'Conc': float(row["Conc. (μmol)"]), 'Dilution': dil, 'Q': total_q, 'n_e': n_e, 'F': F_const}
                fe_res.append(round(eval(st.session_state.hcell_formula, {"__builtins__": {}}, env), 2))
            else:
                env_n = {'C1': float(row["Acid C1 (mM)"]), 'C2': float(row["RE C2 (mM)"]), 'V_acid': st.session_state.acid_vol, 'V_re': st.session_state.re_vol, 'Dilution': dil}
                tn = eval(st.session_state.gde_n_formula, {"__builtins__": {}}, env_n)
                tn_res.append(round(tn, 3))
                env_fe = {'Total_n': tn, 'Q': total_q, 'n_e': n_e, 'F': F_const}
                fe_res.append(round(eval(st.session_state.gde_fe_formula, {"__builtins__": {}}, env_fe), 2))
        except Exception as e:
            fe_res.append("Error")
            if "GDE" in mode: tn_res.append("Error")
    
    if "GDE" in mode: res_df["Total n (μmol)"] = tn_res
    res_df["FE (%)"] = fe_res
    
    st.session_state.res_df = res_df

if 'res_df' in st.session_state:
    st.success("✅ 計算完成！")
    drop_cols = ["選取"]
    if not st.session_state.custom_ne_toggle:
        drop_cols.append("n_e")
    st.dataframe(st.session_state.res_df.drop(columns=drop_cols, errors='ignore'), use_container_width=True)

    def to_pro_excel(df, curr_mode, curr_electrolyte, curr_Q, is_n2, show_ne):
        def apply_subscript(text):
            if pd.isna(text): return ""
            text_str = str(text)
            try:
                f_val = float(text_str)
                if np.isnan(f_val): return ""
                return int(f_val) if f_val.is_integer() else f_val
            except ValueError: pass
            subscript_map = str.maketrans("0123456789", "₀₁₂₃₄₅₆₇₈₉")
            return re.sub(r'([a-zA-Z])(\d+)', lambda m: m.group(1) + m.group(2).translate(subscript_map), text_str)

        wb = Workbook()
        ws = wb.active
        ws.title = "FE_Results"
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        export_data = []
        gas_mode_str = "N2_Gas" if is_n2 else "Ar_Gas"

        for _, row in df.iterrows():
            row_dict = {
                'Cell': f"H-cell({row['Product']})" if "H-cell" in curr_mode else f"GDE_{gas_mode_str}({row['Product']})",
                'Electrolyte': curr_electrolyte, 'Total Coulomb (Q)': curr_Q,
                'Product Type': row['Product']
            }
            if show_ne: row_dict['n_e'] = row['n_e']
            row_dict.update({
                'Catalyst': row['Catalyst'], 'Loading (μl)': row['Loading (μl)'],
                'Dilution Factor': row['Dilution Factor'], 'V vs RHE': row['V vs RHE']
            })
            if "H-cell" in curr_mode:
                row_dict['Total Concentration (μmol)'] = row['Conc. (μmol)']
            else:
                row_dict['Acid C1 (mM)'] = row['Acid C1 (mM)']
                row_dict['RE C2 (mM)'] = row['RE C2 (mM)']
                row_dict['Total Concentration (μmol)'] = row['Total n (μmol)']
            row_dict['Faradaic Efficiency (%)'] = row['FE (%)']
            export_data.append(row_dict)
                
        df_export = pd.DataFrame(export_data)

        cur_row = 1
        for prod, group in df_export.groupby("Product Type"):
            cols = list(group.columns)
            for c_idx, c_name in enumerate(cols):
                cell = ws.cell(row=cur_row, column=c_idx+1, value=apply_subscript(c_name))
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            start_r = cur_row + 1
            for r_idx, r_data in enumerate(group.values.tolist()):
                for c_idx, val in enumerate(r_data):
                    cell = ws.cell(row=start_r+r_idx, column=c_idx+1, value=apply_subscript(val))
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            end_r = start_r + len(group) - 1

            if end_r > start_r:
                merge_indices = [1, 2, 3, 4]
                if show_ne: merge_indices.append(5)
                for col_idx in merge_indices:
                    ws.merge_cells(start_row=start_r, end_row=end_r, start_column=col_idx, end_column=col_idx)

                cat_col_idx = 6 if show_ne else 5
                catalyst_starts = start_r
                current_cat = ws.cell(row=start_r, column=cat_col_idx).value
                for r in range(start_r + 1, end_r + 2):
                    cell_val = ws.cell(row=r, column=cat_col_idx).value if r <= end_r else None
                    if cell_val != current_cat:
                        if (r - 1) > catalyst_starts:
                            ws.merge_cells(start_row=catalyst_starts, end_row=r-1, start_column=cat_col_idx, end_column=cat_col_idx)
                            ws.merge_cells(start_row=catalyst_starts, end_row=r-1, start_column=cat_col_idx+1, end_column=cat_col_idx+1)
                        catalyst_starts = r
                        current_cat = cell_val

            cur_row = end_r + 2

        for col in ws.columns:
            max_length = 0
            column_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2

        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    st.markdown("##### 📁 匯出報表 (可自訂檔名)")
    col_input, col_dl, _ = st.columns([2, 2, 4])
    with col_input:
        custom_excel_name = st.text_input("自訂 Excel 檔名", value=default_excel_name, label_visibility="collapsed", key="excel_name_input")
        excel_filename_final = f"{today_str}_{st.session_state.excel_name_input}.xlsx"
    with col_dl:
        st.download_button("📥 下載 Excel", data=to_pro_excel(st.session_state.res_df, mode, electrolyte, total_q, is_n2_mode, st.session_state.custom_ne_toggle), file_name=excel_filename_final, use_container_width=True)
